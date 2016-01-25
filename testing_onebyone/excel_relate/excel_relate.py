from openpyxl.cell import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
import os
class excel_writer:

    def __init__(self,log_dir, FILE_NAME):
        self.wb = Workbook()
        self.ew = ExcelWriter(workbook=self.wb)
        self.filename = os.path.join(log_dir, FILE_NAME)
        self.ws_num = 0
        self.ws = []

    def create_sheet(self, sheet_title):
        self.ws.append(self.wb.create_sheet(index = self.ws_num, title = sheet_title))
        #self.ws[self.ws_num].title = sheet_title
        self.ws_num += 1

    def set_cell(self, sheet_title, col, row, val):
        ws = self.wb.get_sheet_by_name(sheet_title)
        ws.cell('%s%s' % (get_column_letter(col), row)).value = '%s' % (val)
    #        str(val.decode('utf-8')))

    def save(self):
        self.wb.save(filename=self.filename)

if __name__ == '__main__':
    obj = excel_writer('hehe.xlsx')
    obj.create_sheet('123')
    obj.create_sheet('234')
    obj.set_cell('123', 1, 1, 'haha')
    obj.set_cell('234', 1, 2, 'hehehe')
    obj.save()
